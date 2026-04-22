"""XLSX → Markdown converter.

Each non-empty sheet becomes a tuple of (sheet_name, markdown_table).
The worker assembles the final Markdown sections with `## Лист: <name>` headers.
"""

from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook


def xlsx_to_markdown_sections(file_bytes: bytes) -> list[tuple[str, str]]:
    """
    Parse an XLSX file and return (sheet_name, markdown_table) for each
    non-empty sheet.

    Rules:
    - Empty rows (all cells None) are skipped.
    - The first non-empty row is treated as the table header.
    - Numbers that are whole-valued floats are formatted without the `.0`.
    - None cells are rendered as empty strings.
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sections: list[tuple[str, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = _collect_rows(ws)

        if not rows:
            continue

        sections.append((sheet_name, _rows_to_md_table(rows)))

    wb.close()
    return sections


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _collect_rows(ws) -> list[list[str]]:  # noqa: ANN001
    """Return non-empty rows as lists of formatted cell strings."""
    result: list[list[str]] = []

    for row in ws.iter_rows(values_only=True):
        if not any(cell is not None for cell in row):
            continue

        str_cells = [_format_cell(cell) for cell in row]

        # Drop trailing empty cells to keep tables clean.
        while str_cells and str_cells[-1] == "":
            str_cells.pop()

        if str_cells:
            result.append(str_cells)

    return result


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    # datetime must be checked before date (datetime is a subclass of date).
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    # Whole-number floats: show as integer (5800.0 → "5800").
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _rows_to_md_table(rows: list[list[str]]) -> str:
    """Build a Markdown table. First row is the header."""
    if not rows:
        return ""

    col_count = max(len(row) for row in rows)

    header = _pad_row(rows[0], col_count)
    md_lines: list[str] = [
        "| " + " | ".join(header) + " |",
        "|" + " | ".join(["---"] * col_count) + "|",
    ]

    for row in rows[1:]:
        md_lines.append("| " + " | ".join(_pad_row(row, col_count)) + " |")

    return "\n".join(md_lines)


def _pad_row(row: list[str], width: int) -> list[str]:
    return (row + [""] * width)[:width]
